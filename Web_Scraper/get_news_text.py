import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os
import re

# Load the workbook
wb = load_workbook('edweek_news.xlsx')
# Define the directory to save the text files
directory = 'edu_news'

# Create the directory if it doesn't exist
if not os.path.exists(directory):
    os.makedirs(directory)

# Iterate over each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Iterate over each row in the sheet, skipping the header
    for row in ws.iter_rows(min_row=2, values_only=True):
        title, link = row

        # Request the news content
        response = requests.get(link)
        soup = BeautifulSoup(response.text, 'html.parser')

        # Extract the content using the specified selector
        content_elements = soup.select('.a-text p, .a-text > h2')
        content = '\n'.join([element.get_text(strip=True) for element in content_elements])

        # Clean the title to create a valid filename
        # Define a regular expression to match invalid filename characters
        invalid_chars = r'[\\/:"*?<>|]'

        # Clean the title
        filename = f"{re.sub(invalid_chars, '-', title[:250]).strip()}.txt"
        filepath = os.path.join(directory, filename)


        # Save the content to a .txt file
        with open(filepath, 'w', encoding='utf-8') as file:
            file.write(content)
        # Print a confirmation message
        print(f"Saved: {filename}")

print("News content has been saved to .txt files.")
