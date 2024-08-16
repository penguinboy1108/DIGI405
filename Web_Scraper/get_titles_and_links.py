import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Initialize a workbook
wb = Workbook()

for i in range(5):
    response = requests.get(f'https://www.edweek.org/search?q=ai+student&s=0&p={i + 1}')
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extract titles and links
    articles = soup.select('a.m-promo__title')
    titles = [article.get_text(strip=True) for article in articles]
    links = [article['href'] for article in articles]

    # Create a new sheet for each page
    ws = wb.create_sheet(title=f'Page_{i + 1}')

    # Add headers
    ws.append(['Title', 'Link'])

    # Add the data
    for title, link in zip(titles, links):
        ws.append([title, link])

# Remove the default sheet created by openpyxl
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the workbook
wb.save('edweek_news.xlsx')