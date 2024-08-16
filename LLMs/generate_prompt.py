import logging
import requests
import json
import openpyxl
import re
from setting import API_KEY,modelList
import os

logging.basicConfig(level=logging.INFO)

OPENROUTER_API_KEY = API_KEY


def query_llm(prompt: str,  # prompt to send to LLM
              model: str,  # model name e.g. google/gemma-2-9b-it:free
              system_prompt: str = None,  # system prompt to send to LLM
              max_tokens: int = 2048,  # maximum number of tokens to generate (includes prompt tokens)
              response_format: str = None,  # response format: json or None
              temperature: float = None  # temperature for sampling
              ) -> requests.models.Response:
    """ Query LLM with prompt """

    api_url = "https://openrouter.ai/api/v1/chat/completions"
    if OPENROUTER_API_KEY is None:
        logging.error("OPENROUTER_API_KEY not set. Not querying llm.")
        return None
    api_key = OPENROUTER_API_KEY

    if prompt.strip() == '':
        logging.error('No prompt provided. Not querying llm.')
        return None

    messages = []
    if system_prompt is not None and system_prompt.strip() != '':
        messages.append({"role": "system", "content": system_prompt})
    messages.append({"role": "user", "content": prompt})

    request_data = {
        "model": model,
        "messages": messages,
        'max_tokens': max_tokens
    }

    if temperature is not None:
        request_data['temperature'] = temperature

    if response_format == "json":
        request_data['response_format'] = {"type": "json_object"}

    text = None

    try:
        response = requests.post(
            url=api_url,
            headers={
                "Authorization": f"Bearer {api_key}",
            },
            data=json.dumps(request_data)
        )
        response.raise_for_status()
        text = response.json()['choices'][0]['message']['content']
    except requests.exceptions.RequestException as e:
        logging.error(f"Error querying LLM: {e}")
    except Exception as e:
        logging.error(f"Error querying LLM: {e}")

    return text


role_list = ['teachers','students'][::-1]
task_list = ['teaching','studying'][::-1]

for j in range(len(role_list)):
    wb = openpyxl.Workbook()
    role = role_list[j]  # 可替换为任何角色，例如 "student"
    task = task_list[j]  # 可替换为任务或主题，例如 "learning"
    for i, model in enumerate(modelList):
        print(i, model)
        # 定义可输入的选项
        tea_exam = 'AI will help other complete their homework, making it difficult to verify students learning outcomes.'
        stu_exam = 'AI will help me do my homework and complete the difficult or meaningless homework assigned by the teacher'
        # 定义system_prompt
        system_prompt = f'''
        From the perspective of {role}
        '''

        # 定义prompt
        prompt = f'''Generate 20 Answers in the form of a statement from the article title related to the following topics:
        How AI affects your {task}
        It may have a positive or negative impact on your {task}. For example, AI provides new methods and can generate more questions to help your {task}. 
        '''

        # 定义context_prompt
        context_prompt = f'''
        Generate text from the perspective of {role}s, no less than 1000 words
        '''
        max_tokens = 2048
        MAX_FILENAME_LENGTH = 64
        # 生成文本
        text = query_llm(prompt, model, system_prompt, max_tokens)
        cleaned_text = re.sub(r'^\d+\.\s*', '', text, flags=re.MULTILINE)
        lines = [line for line in cleaned_text.split('\n')[1:] if line.strip()]

        print(lines)

        output_dir = role+'_gen_texts'
        os.makedirs(output_dir, exist_ok=True)

        # 为每个模型创建一个新的工作表
        ws = wb.create_sheet(title=f"Model_{i + 1}")

        # 生成文本并存储在文件和Excel中
        for line_idx, line in enumerate(lines, start=1):
            gen_texts = query_llm(line + context_prompt, model, context_prompt, max_tokens)

            sanitized_filename = ''.join(char for char in line if char.isalnum() or char.isspace()).replace(' ', '_')
            # 如果文件名过长，进行截断
            if len(sanitized_filename) > MAX_FILENAME_LENGTH - 4:  # 留出 ".txt" 的4个字符空间
                sanitized_filename = sanitized_filename[:MAX_FILENAME_LENGTH - 4]
            file_path = os.path.join(output_dir, f"{sanitized_filename}.txt")

            # 将生成的文本写入 .txt 文件
            with open(file_path, 'w', encoding='utf-8') as file:
                file.write(gen_texts)
            print(f"Saved context to {file_path}")

            # 在Excel中填入编号和生成的文本
            ws.cell(row=line_idx, column=1, value=line_idx)  # 第一列是编号
            ws.cell(row=line_idx, column=2, value=line)  # 第二列是生成的文本内容

    # 移除默认创建的第一个空白工作表
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 保存 Excel 文件
    wb.save(role+"_output.xlsx")